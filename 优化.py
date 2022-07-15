import datetime
import requests
import bs4
import re
import openpyxl

header = {
    'user-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/102.0.5005.61 Safari/537.36'}


def creating_urls(page_number: int):
    url_dic = {}
    for i in range(page_number):
        url_dic[
            i] = f'http://77.push2.eastmoney.com/api/qt/clist/get?cb=jQuery11240010983169070974963_1655099439009&pn={i + 1}&pz=20&po=1&np=1&ut=bd1d9ddb04089700cf9c27f6f7426281&fltt=2&invt=2&wbp2u=|0|0|0|web&fid=f3&fs=m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23,m:0+t:81+s:2048&fields=f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f12,f13,f14,f15,f16,f17,f18,f20,f21,f23,f24,f25,f22,f11,f62,f128,f136,f115,f152&_=1655099439081'
        url_list = list(url_dic.values())
    j = 1;
    bs4_objects = []
    for url in url_list:
        target = requests.get(url)
        object = bs4.BeautifulSoup(target.text, "lxml")
        # object=object.find_all("p",attrs={"class":"f14:"})
        bs4_objects.append(object)
        print(f"已下载第{j}页，{url}")
        j += 1
    return bs4_objects


def save_excel(name: list,number:list,percen:list, date: str):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = date
    i = 0; j = 0; k = 0
    for rows_value in name:
        sheet.cell(row=1, column=1 + i, value=rows_value)
        i += 1
    for rows_value in number:
        sheet.cell(row=2, column=1 + j, value=rows_value)
        j += 1
    for rows_value in percen:
        sheet.cell(row=3, column=1 + k, value=rows_value)
        k += 1

    try:
        workbook.save("share.xlsx")
        return {'status': True}
    except Exception as e:
        return {'status': False, 'error': '原因如下：' + str(e)}


def get_stock_data(text):
    com = re.compile('"f2":(?P<end>.+?),.*?"f3":(?P<percentage>.+?),.*?"f12":"(?P<number>.+?)",.*?"f14":"(?P<name>.+?)"'
                     ',.*?"f17":(?P<start>.+?),', re.S)
    ret = com.finditer(text)

    for i in ret:
        yield {
            'number': i.group('number'),
            'name': i.group('name'),
            'start': i.group('start'),
            'end': i.group('end'),
            'percentage': i.group('percentage')
        }


# 生成数据类型
name_list = []
number_list = []
percentage_list = []
end_list = []
start_list = []

bs_rawdoc = creating_urls(int(input("生成页数")))
print(bs_rawdoc)
data = get_stock_data(str(bs_rawdoc))
date = datetime.date.today()
for j in data:
    number = j.get('number')
    name = j.get('name')
    start = j.get('start')
    end = j.get('end')
    percentage = j.get('percentage')
    if start == '"-"':
        start, end, percentage = '0', '0', '0'

#if float(percentage) > 9.4 and float(percentage) < 10.4:
    else:
        name_list.append(name)
        number_list.append(number)
        percentage_list.append(percentage)
        end_list.append(end)
        start_list.append(start)
print(f'''{name_list}
{number_list}
{percentage_list}''')
result = save_excel(name_list,number_list,percentage_list, str(date))
